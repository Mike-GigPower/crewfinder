"""
Gig Power post-show timesheet importer — workbook (.xlsx) parsing.

A booking's timesheet is one .xlsx workbook: each CALL is its own tab (e.g.
"Fri 1800 (Show Crew)"), plus support tabs we ignore (DATA LIST, Schedule, …).
Every call tab has a header on row 16 and crew from row 17 down, keyed by EIN
(column K) — which matches SmartStaff users.ein exactly, so no fuzzy matching.

We read the workbook's CACHED computed values (Google formulas don't survive the
xlsx export, but their last-computed values do). Clock times come through as
datetimes with sub-second float dust from the sheet's CEILING rounding (e.g.
22:59:59.712 means 23:00), so on/off are rounded to the nearest minute.

The column map and per-cell parsing now live in timesheet_common, shared with the
live Google Sheets reader so the two import sources can't drift. This module keeps
only the openpyxl-specific bits: detecting a call tab and reading J2 for the
tab->call auto-map tie-break.
"""

from io import BytesIO

from timesheet_common import (
    HEADER_ROW, FIRST_DATA_ROW, _LABELS, header_map, parse_crew_row,
)


def _is_call_tab(ws):
    """A call tab has STATUS / Start Time / EIN / LAST NAME on the header row."""
    labels = set()
    for c in range(1, ws.max_column + 1):
        val = ws.cell(HEADER_ROW, c).value
        if isinstance(val, str):
            labels.add(val.strip().lower())
    return {"ein", "start time", "last name"}.issubset(labels)


def _header_cells(ws):
    """The header row as a plain list (column A first), for timesheet_common.header_map."""
    return [ws.cell(HEADER_ROW, c).value for c in range(1, ws.max_column + 1)]


def _call_time_iso(ws):
    """The scheduled GIG Call Time in J2 (for auto-mapping tab -> call)."""
    from datetime import datetime
    v = ws.cell(2, 10).value  # J2
    if isinstance(v, datetime):
        return v.isoformat()
    return None


def parse_timesheet_workbook(source):
    """Parse a timesheet workbook (path or bytes) into per-call-tab time rows.

    Returns:
      {
        "tabs": [
          { "tab_name", "call_time" (iso|None),
            "rows": [ {ein, lastname, firstname, status, on, off,
                       break, break_night, late, note, no_show} ] }
        ],
        "skipped_tabs": [name, ...]   # non-call tabs
      }
    """
    import openpyxl  # imported lazily so the rest of the app doesn't need it

    if isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(BytesIO(source), data_only=True)
    else:
        wb = openpyxl.load_workbook(source, data_only=True)

    tabs, skipped = [], []

    for ws in wb.worksheets:
        if not _is_call_tab(ws):
            skipped.append(ws.title)
            continue

        col = header_map(_header_cells(ws))
        rows = []

        for r in range(FIRST_DATA_ROW, ws.max_row + 1):
            def get_cell(c, _r=r):
                return ws.cell(_r, c).value
            row = parse_crew_row(get_cell, col)
            if row is None:
                continue  # past the crew block
            rows.append(row)

        tabs.append({
            "tab_name":  ws.title,
            "call_time": _call_time_iso(ws),
            "rows":      rows,
        })

    wb.close()
    return {"tabs": tabs, "skipped_tabs": skipped}
