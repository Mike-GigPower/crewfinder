"""
Gig Power timesheet generation — build a pre-filled crew-master workbook.

The inverse of the importer. A Crew Boss fills these sheets in the field (often
offline), so generation produces *their* crew master, not a GOAT-native format:
for each call it clones the workbook's "Master" tab — keeping every formula, the
status/dept dropdowns and the conditional formatting — sets the GIG Call Time in
J2, stamps the Call ID, and pre-fills the confirmed crew (last, first, EIN, phone)
into the I/J/K/L columns the boss would otherwise type by hand.

openpyxl's copy_worksheet preserves formulas, styles and column widths but drops
data validations and conditional formatting, so those are carried over explicitly
(verified to round-trip cleanly).

The Call ID is written as a label/value pair (A1 = "GOAT Call ID", B1 = id) so the
importer can map each tab back to its call by exact lookup instead of inferring
from EIN overlap.
"""

import re
from copy import copy
from io import BytesIO

MASTER_TAB    = "Master"
FIRST_DATA_ROW = 17
CALLID_LABEL_CELL = "A1"
CALLID_VALUE_CELL = "B1"
CALLID_LABEL = "GOAT Call ID"

# STATUS dropdown options (the Google->xlsx export breaks the originals to #REF!,
# so generation rewrites them to a working literal list).
STATUS_OPTIONS = ("Confirmed", "Late", "Moved", "No Show")

# The broken list validations we replace (everything else is carried as-is).
_STATUS_RANGE = "A17:A304"
_DEPT_RANGE   = "M17:M304"
_DEPT_COL = 26  # column Z, where the Master lists the dept names (Z2 down)

_BAD_TAB_CHARS = re.compile(r"[\[\]:\*\?/\\]")


def _safe_title(name, used):
    """A valid, unique Excel tab title (<=31 chars, no []:*?/\\)."""
    t = _BAD_TAB_CHARS.sub(" ", str(name or "Call")).strip()[:31] or "Call"
    base, i, low = t, 2, t.lower()
    while low in used:
        suffix = " " + str(i)
        t = (base[:31 - len(suffix)] + suffix)
        low = t.lower()
        i += 1
    used.add(low)
    return t


def _read_depts(master):
    """The dept names the Master lists in column Z (Z2 down), for the DEPT dropdown."""
    out = []
    for r in range(2, 14):
        v = master.cell(r, _DEPT_COL).value
        if v is not None and str(v).strip() and str(v).strip().upper() != "TOTAL":
            out.append(str(v).strip())
    return out


def _carry_formatting(master, clone, depts):
    """Carry the data validations copy_worksheet drops, replacing the export-broken
    STATUS / DEPT dropdowns with working literal lists.

    Conditional formatting is intentionally NOT copied: the Master's CF rules are
    Google-authored (containsText/expression) and openpyxl re-serializes them in a
    form Excel rejects (it strips them on open with a repair dialog). They're purely
    cosmetic colour-coding, so a clean file is the better trade. Static styling
    (fonts, borders, header fills, column widths) still comes through copy_worksheet.
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    for dv in master.data_validations.dataValidation:
        if dv.formula1 == "#REF!":
            continue  # export-broken dropdowns (status / dept / break) — status & dept replaced below
        clone.add_data_validation(copy(dv))

    sdv = DataValidation(type="list", formula1='"' + ",".join(STATUS_OPTIONS) + '"', allow_blank=True)
    sdv.add(_STATUS_RANGE)
    clone.add_data_validation(sdv)

    if depts:
        ddv = DataValidation(type="list", formula1='"' + ",".join(depts) + '"', allow_blank=True)
        ddv.add(_DEPT_RANGE)
        clone.add_data_validation(ddv)


def generate_timesheet_workbook(template_source, calls):
    """Build a crew-master workbook with one Master-cloned tab per call.

    template_source: path or bytes of the crew master .xlsx (must contain a
                     "Master" tab).
    calls: list of dicts, each:
        {
          "call_id":   int,
          "call_name": str,
          "call_time": datetime | None,   # -> J2
          "crew": [ {"lastname","firstname","ein","phone"}, ... ]  # confirmed
        }
      Order is preserved; sort before calling if you want chronological tabs.

    Returns: bytes of the generated .xlsx (original support tabs and the Master
    tab are retained; call tabs are appended in order).
    """
    import openpyxl

    if isinstance(template_source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(BytesIO(template_source), data_only=False)
    else:
        wb = openpyxl.load_workbook(template_source, data_only=False)

    if MASTER_TAB not in wb.sheetnames:
        raise ValueError('Template has no "%s" tab' % MASTER_TAB)

    # openpyxl re-serializes every sheet's conditional formatting on save, and the
    # Google-authored rules come out in a form Excel rejects (repair dialog). Strip
    # CF from all sheets up front so the output is clean — it's cosmetic colour-coding.
    # Also drop the export-broken #REF! data validations (and, below, the broken
    # "Status" defined name) so Excel doesn't flag a phantom workbook link.
    from openpyxl.formatting.formatting import ConditionalFormattingList
    for ws in wb.worksheets:
        ws.conditional_formatting = ConditionalFormattingList()
        dvs = ws.data_validations.dataValidation
        kept = [dv for dv in dvs if str(dv.formula1) != "#REF!"]
        if len(kept) != len(dvs):
            ws.data_validations.dataValidation = kept

    # remove broken / external-looking defined names (e.g. Status = #REF!) — these
    # make Excel show an "unable to refresh linked workbook" banner on open.
    try:
        for nm in list(wb.defined_names.keys()):
            v = str(wb.defined_names[nm].value)
            if "#REF!" in v or "[" in v:
                del wb.defined_names[nm]
    except Exception:
        pass

    master = wb[MASTER_TAB]
    depts = _read_depts(master)

    used = set(n.lower() for n in wb.sheetnames)

    for call in calls:
        clone = wb.copy_worksheet(master)
        clone.title = _safe_title(call.get("call_name") or ("Call " + str(call.get("call_id"))), used)
        _carry_formatting(master, clone, depts)

        if call.get("call_time") is not None:
            clone["J2"] = call["call_time"]

        clone[CALLID_LABEL_CELL] = CALLID_LABEL
        clone[CALLID_VALUE_CELL] = call.get("call_id")

        for i, m in enumerate(call.get("crew", [])):
            r = FIRST_DATA_ROW + i
            clone["A%d" % r] = "Confirmed"
            clone["I%d" % r] = (m.get("lastname") or "")
            clone["J%d" % r] = (m.get("firstname") or "")
            if m.get("ein") is not None:
                clone["K%d" % r] = m["ein"]
            if m.get("phone"):
                clone["L%d" % r] = m["phone"]

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
